import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")


def _apply_header(ws: Worksheet, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def build_excel(data: dict, from_date: date, to_date: date) -> io.BytesIO:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = "SupportOps Report"
    ws_summary["A2"] = f"Period: {from_date} to {to_date}"
    ws_summary["A4"] = "Total Request Volume"
    request_volume = data.get("request_volume", {})
    ws_summary["B4"] = request_volume.get("total", 0) if isinstance(request_volume, dict) else request_volume
    sla = data.get("sla_health", {})
    ws_summary["A5"] = "SLA Compliance Rate (%)"
    ws_summary["B5"] = sla.get("compliance_rate_pct", sla.get("compliance_rate", 0))

    ws_status = wb.create_sheet("Status Breakdown")
    _apply_header(ws_status, ["Status", "Count"])
    status_breakdown = data.get("status_breakdown", {})
    if isinstance(status_breakdown, dict):
        for i, (status, count) in enumerate(status_breakdown.items(), 2):
            ws_status.cell(row=i, column=1, value=status)
            ws_status.cell(row=i, column=2, value=count)
    else:
        for i, row in enumerate(status_breakdown, 2):
            ws_status.cell(row=i, column=1, value=row.get("status", ""))
            ws_status.cell(row=i, column=2, value=row.get("count", 0))

    ws_sla = wb.create_sheet("SLA Health")
    _apply_header(ws_sla, ["Metric", "Value"])
    ws_sla.cell(row=2, column=1, value="Total SLA Records")
    ws_sla.cell(row=2, column=2, value=sla.get("total", 0))
    ws_sla.cell(row=3, column=1, value="Breached")
    ws_sla.cell(row=3, column=2, value=sla.get("breached", 0))
    ws_sla.cell(row=4, column=1, value="Compliance Rate (%)")
    ws_sla.cell(row=4, column=2, value=sla.get("compliance_rate_pct", sla.get("compliance_rate", 0)))

    ws_team = wb.create_sheet("Team Performance")
    _apply_header(ws_team, ["Technician", "Assigned", "Resolved"])
    for i, row in enumerate(data.get("team_performance", []), 2):
        ws_team.cell(row=i, column=1, value=row.get("technician", ""))
        ws_team.cell(row=i, column=2, value=row.get("assigned", 0))
        ws_team.cell(row=i, column=3, value=row.get("resolved", 0))

    ws_svc = wb.create_sheet("Service Types")
    _apply_header(ws_svc, ["Service Type", "Count"])
    for i, row in enumerate(data.get("service_type_breakdown", []), 2):
        ws_svc.cell(row=i, column=1, value=row.get("service_type", ""))
        ws_svc.cell(row=i, column=2, value=row.get("count", 0))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
